from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from datetime import datetime

wb = Workbook()
ws = wb.active
ws.title = "Suivi tâches"

# Couleurs
orange = "F4B183"
green = "C6EFCE"
gray = "D9EAD3"
dark_gray = "BFBFBF"
header_blue = "1F4E78"
white = "FFFFFF"
light_yellow = "FFF2CC"
thin_gray = "D9D9D9"

# Légende
ws["A1"] = "Légende"
ws["A1"].font = Font(bold=True, size=13)

legend = [
    ("A2", "X", orange, "Tâche assignée au membre"),
    ("A3", "OK", green, "Tâche membre terminée"),
    ("A4", "Contrôleur vert", green, "Au moins un membre a mis OK"),
    ("A5", "Ligne barrée/grisée", dark_gray, "Statut = Terminée"),
]

for cell, label, fill, desc in legend:
    ws[cell] = label
    ws[cell].fill = PatternFill("solid", fgColor=fill)
    ws[cell].font = Font(bold=True)
    ws[cell.replace("A", "B")] = desc

# En-têtes
headers = [
    "Tâche",
    "Deadline",
    "Priorité",
    "Statut",
    "Contrôleur",
    "Alice",
    "Bob",
    "Charlie",
    "Diana",
]

start_row = 7
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=start_row, column=col, value=header)
    cell.fill = PatternFill("solid", fgColor=header_blue)
    cell.font = Font(color=white, bold=True)
    cell.alignment = Alignment(horizontal="center")

priority_rank = {"Haute": 1, "Moyenne": 2, "Basse": 3}

tasks = [
    ["Finaliser cahier des charges", "24/06/2026", "Haute", "En cours", "Marie", "X", "OK", "", ""],
    ["Préparer maquette interface", "27/06/2026", "Moyenne", "À faire", "Marie", "", "X", "X", ""],
    ["Valider budget projet", "23/06/2026", "Haute", "En cours", "Marie", "OK", "", "", "X"],
    ["Créer base de données", "02/07/2026", "Haute", "À faire", "Marie", "", "X", "", "X"],
    ["Rédiger documentation", "05/07/2026", "Basse", "À faire", "Marie", "X", "", "X", ""],
    ["Tester module paiement", "29/06/2026", "Haute", "En cours", "Marie", "", "OK", "X", ""],
    ["Former équipe support", "08/07/2026", "Moyenne", "À faire", "Marie", "X", "", "", "X"],
    ["Livrer version pilote", "10/07/2026", "Haute", "À faire", "Marie", "X", "X", "X", "X"],
    ["Nettoyer backlog", "01/07/2026", "Basse", "Terminée", "Marie", "OK", "OK", "", ""],
]

# Tri initial deadline puis priorité
tasks.sort(
    key=lambda x: (
        datetime.strptime(x[1], "%d/%m/%Y"),
        priority_rank[x[2]]
    )
)

for r, task in enumerate(tasks, start_row + 1):
    for c, value in enumerate(task, 1):
        cell = ws.cell(row=r, column=c, value=value)
        if c == 2:
            cell.number_format = "DD/MM/YYYY"
            cell.value = datetime.strptime(value, "%d/%m/%Y")
        cell.alignment = Alignment(vertical="center")

last_row = start_row + len(tasks)
last_col = len(headers)

# Tableau Excel
table_ref = f"A{start_row}:{get_column_letter(last_col)}{last_row}"
tab = Table(displayName="TableauTaches", ref=table_ref)
style = TableStyleInfo(
    name="TableStyleMedium2",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False,
)
tab.tableStyleInfo = style
ws.add_table(tab)

# Validations listes
dv_priority = DataValidation(type="list", formula1='"Haute,Moyenne,Basse"', allow_blank=False)
dv_status = DataValidation(type="list", formula1='"À faire,En cours,Terminée"', allow_blank=False)
ws.add_data_validation(dv_priority)
ws.add_data_validation(dv_status)
dv_priority.add(f"C{start_row+1}:C1000")
dv_status.add(f"D{start_row+1}:D1000")

# Mise en forme conditionnelle
member_range = f"F{start_row+1}:I1000"

# X orange
ws.conditional_formatting.add(
    member_range,
    CellIsRule(
        operator="equal",
        formula=['"X"'],
        fill=PatternFill("solid", fgColor=orange)
    )
)

# OK vert
ws.conditional_formatting.add(
    member_range,
    CellIsRule(
        operator="equal",
        formula=['"OK"'],
        fill=PatternFill("solid", fgColor=green)
    )
)

# Contrôleur vert si au moins un OK sur ligne
for row in range(start_row + 1, 1001):
    ws.conditional_formatting.add(
        f"E{row}",
        FormulaRule(
            formula=[f'COUNTIF($F{row}:$I{row},"OK")>0'],
            fill=PatternFill("solid", fgColor=green)
        )
    )

# Ligne grisée + barrée si statut Terminée
for row in range(start_row + 1, 1001):
    ws.conditional_formatting.add(
        f"A{row}:I{row}",
        FormulaRule(
            formula=[f'$D{row}="Terminée"'],
            fill=PatternFill("solid", fgColor=dark_gray),
            font=Font(strike=True, color="666666")
        )
    )

# Bordures
side = Side(style="thin", color=thin_gray)
for row in ws.iter_rows(min_row=start_row, max_row=last_row, min_col=1, max_col=last_col):
    for cell in row:
        cell.border = Border(left=side, right=side, top=side, bottom=side)

# Largeurs colonnes
widths = {
    "A": 32,
    "B": 14,
    "C": 12,
    "D": 14,
    "E": 16,
    "F": 12,
    "G": 12,
    "H": 12,
    "I": 12,
}
for col, width in widths.items():
    ws.column_dimensions[col].width = width

# Figer en-tête
ws.freeze_panes = "A8"

# Filtre déjà dans tableau. Utilisateur peut filtrer Statut != Terminée.
# Masquage auto sans macro impossible en xlsx pur.

# Alignement centre colonnes courtes
for col in range(2, last_col + 1):
    for row in range(start_row + 1, 1001):
        ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")

# Note
ws["K7"] = "Note"
ws["K8"] = "Tri auto/masquage auto complet nécessite macro VBA (.xlsm)."
ws["K9"] = "Dans ce .xlsx : tri initial + filtres + mise en forme conditionnelle."

wb.save("suivi_taches.xlsx")
print("Fichier créé : suivi_taches.xlsx")